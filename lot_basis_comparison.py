import argparse
import csv
import html
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

SCHWAB_LOTS_DIR = Path("SchwabLots")
QUICKEN_EXPORTS_DIR = Path("QuickenExports")
QUICKEN_LOTS_RE = re.compile(r"^QuickenLots_(\d{4}-\d{2}-\d{2})\.xlsx$")

SYMBOL_COL = "Symbol"
OPEN_DATE_COL = "Open Date"
QUANTITY_COL = "Quantity"
COST_BASIS_COL = "Cost Basis"
MARKET_VALUE_COL = "Market Value"
CASH_SYMBOL = "CASH"
CASH_LOT_KEY = "__cash__"


@dataclass
class Lot:
    """Represents a single lot of shares within a security."""

    name: str
    quote: float = 0.0
    shares: float = 0.0
    market_value: float = 0.0
    cost_basis: float = 0.0


@dataclass
class Security:
    """Represents a Quicken security with totals and individual lots."""

    name: str
    ticker: str
    quote: float = 0.0
    shares: float = 0.0
    market_value: float = 0.0
    cost_basis: float = 0.0
    lots: list[Lot] = field(default_factory=list)


@dataclass
class AccountSecurities:
    securities: dict[str, Security]
    account_name: Optional[str]


@dataclass
class SchwabLots:
    path: Path
    account_suffix: Optional[str]
    lots_by_symbol: dict[str, dict[str, dict[str, float]]]


@dataclass
class LotComparison:
    """Result of comparing a lot between Schwab and Quicken."""

    date: str
    schwab_shares: float = 0.0
    quicken_shares: float = 0.0
    schwab_cost_basis: float = 0.0
    quicken_cost_basis: float = 0.0
    status: str = ""

    def __str__(self) -> str:
        if self.status == "match":
            return f"  OK {self.date}: {self.schwab_shares:.4f} shares @ ${self.schwab_cost_basis:.2f}"
        if self.status == "missing_in_quicken":
            return (
                f"  MISSING IN QUICKEN - {self.date}: "
                f"Schwab has {self.schwab_shares:.4f} shares @ ${self.schwab_cost_basis:.2f}"
            )
        if self.status == "missing_in_schwab":
            return (
                f"  MISSING IN SCHWAB - {self.date}: "
                f"Quicken has {self.quicken_shares:.4f} shares @ ${self.quicken_cost_basis:.2f}"
            )

        shares_diff = (
            f" shares: {self.schwab_shares:.4f} vs {self.quicken_shares:.4f}"
            if abs(self.schwab_shares - self.quicken_shares) > 0.01
            else ""
        )
        cost_diff = (
            f" cost: ${self.schwab_cost_basis:.2f} vs ${self.quicken_cost_basis:.2f}"
            if abs(self.schwab_cost_basis - self.quicken_cost_basis) > 0.01
            else ""
        )
        return f"  MISMATCH - {self.date}:{shares_diff}{cost_diff}"


@dataclass
class SymbolComparison:
    symbol: str
    name: str
    schwab_shares: float
    schwab_cost_basis: float
    quicken_shares: float
    quicken_cost_basis: float
    schwab_lots: int
    quicken_lots: int
    market_value: float
    lot_comparisons: list[LotComparison] = field(default_factory=list)


def parse_number(value) -> float:
    """Convert Schwab/Quicken formatted numbers to floats."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text or text == "*":
        return 0.0

    text = text.replace("$", "").replace(",", "").replace("*", "").strip()
    text = re.sub(r"\s+\d+$", "", text)

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()
    if text.endswith("-"):
        negative = True
        text = text[:-1].strip()

    try:
        parsed = float(text)
    except ValueError:
        return 0.0
    return -parsed if negative else parsed


def normalize_date(value) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{value.year}"

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            return f"{parsed.month}/{parsed.day}/{parsed.year}"
        except ValueError:
            pass

    parts = text.split("/")
    if len(parts) == 3:
        try:
            year = int(parts[2])
            if year < 100:
                year += 2000
            return f"{int(parts[0])}/{int(parts[1])}/{year}"
        except ValueError:
            return None

    return None


def sortable_date(date_text: str) -> tuple[int, int, int]:
    month, day, year = (int(part) for part in date_text.split("/"))
    return year, month, day


def extract_account_suffix_from_title(title_line: str) -> Optional[str]:
    match = re.search(r"for\s+\.\.\.(\d{3,4})\b", (title_line or "").strip(), re.IGNORECASE)
    return match.group(1) if match else None


def find_schwab_lot_file() -> Path:
    if not SCHWAB_LOTS_DIR.exists():
        raise SystemExit(f"ERROR: Schwab lots directory not found: {SCHWAB_LOTS_DIR.resolve()}")

    paths = sorted(SCHWAB_LOTS_DIR.glob("*.csv"))
    if not paths:
        raise SystemExit(f"ERROR: No Schwab lots CSV files found in {SCHWAB_LOTS_DIR.resolve()}")
    if len(paths) > 1:
        listed = "\n".join(f"  - {path}" for path in paths)
        raise SystemExit(
            "ERROR: Expected exactly one combined Schwab lots CSV in "
            f"{SCHWAB_LOTS_DIR.resolve()}, found {len(paths)}:\n{listed}"
        )
    return paths[0]


def find_latest_quicken_lots_file() -> Path:
    if not QUICKEN_EXPORTS_DIR.exists():
        raise SystemExit(f"ERROR: Quicken exports directory not found: {QUICKEN_EXPORTS_DIR.resolve()}")

    dated_paths: list[tuple[datetime, Path]] = []
    for path in QUICKEN_EXPORTS_DIR.glob("QuickenLots_*.xlsx"):
        match = QUICKEN_LOTS_RE.match(path.name)
        if not match:
            continue
        dated_paths.append((datetime.strptime(match.group(1), "%Y-%m-%d"), path))

    if not dated_paths:
        raise SystemExit(
            "ERROR: No Quicken lots workbook found. Expected files like "
            f"{QUICKEN_EXPORTS_DIR / 'QuickenLots_YYYY-MM-DD.xlsx'}"
        )

    return max(dated_paths, key=lambda item: item[0])[1]


def find_header_row(rows: list[list[str]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows):
        normalized = [cell.strip() for cell in row]
        if SYMBOL_COL in normalized and OPEN_DATE_COL in normalized:
            missing = [
                column
                for column in (SYMBOL_COL, OPEN_DATE_COL, QUANTITY_COL, COST_BASIS_COL, MARKET_VALUE_COL)
                if column not in normalized
            ]
            if missing:
                raise SystemExit(
                    "ERROR: Schwab lots file is missing required column(s): "
                    + ", ".join(missing)
                )
            return index, normalized

    raise SystemExit(
        "ERROR: Could not find the combined Schwab lots header row with "
        f"{SYMBOL_COL!r} and {OPEN_DATE_COL!r}."
    )


def add_lot(
    lots: dict[str, dict[str, float]],
    date_text: str,
    shares: float,
    cost: float,
    market_value: float = 0.0,
) -> None:
    if date_text not in lots:
        lots[date_text] = {"shares": 0.0, "cost_basis": 0.0, "market_value": 0.0}
    lots[date_text]["shares"] += shares
    lots[date_text]["cost_basis"] += cost
    lots[date_text]["market_value"] += market_value


def load_schwab_lots(path: Path) -> SchwabLots:
    text = path.read_text(encoding="utf-8-sig")
    rows = list(csv.reader(text.splitlines()))
    if not rows:
        raise SystemExit(f"ERROR: Schwab lots file is empty: {path.resolve()}")

    title_line = rows[0][0] if rows[0] else ""
    account_suffix = extract_account_suffix_from_title(title_line)
    header_index, headers = find_header_row(rows)

    lots_by_symbol: dict[str, dict[str, dict[str, float]]] = {}
    for row in rows[header_index + 1 :]:
        if not any(cell.strip() for cell in row):
            continue
        values = dict(zip(headers, row))
        symbol = values.get(SYMBOL_COL, "").strip().upper()
        open_date = values.get(OPEN_DATE_COL, "").strip()
        quantity = values.get(QUANTITY_COL, "").strip()

        if not symbol:
            continue
        if symbol == CASH_SYMBOL:
            cost = parse_number(values.get(COST_BASIS_COL, ""))
            market_value = parse_number(values.get(MARKET_VALUE_COL, "")) or cost
            add_lot(lots_by_symbol.setdefault(CASH_SYMBOL, {}), CASH_LOT_KEY, 0.0, cost, market_value)
            continue
        if not open_date or open_date.lower() == "total" or not quantity:
            continue

        date_text = normalize_date(open_date)
        if not date_text:
            continue

        shares = parse_number(quantity)
        cost = parse_number(values.get(COST_BASIS_COL, ""))
        market_value = parse_number(values.get(MARKET_VALUE_COL, ""))
        if abs(shares) < 1e-9:
            continue

        add_lot(lots_by_symbol.setdefault(symbol, {}), date_text, shares, cost, market_value)

    if not lots_by_symbol:
        raise SystemExit(f"ERROR: No stock lots found in combined Schwab file: {path.resolve()}")

    return SchwabLots(path=path, account_suffix=account_suffix, lots_by_symbol=lots_by_symbol)


def parse_quicken_lots(path: Path, account_suffix: Optional[str]) -> AccountSecurities:
    """
    Parse the latest Quicken lots workbook and return securities for the account
    matching the Schwab masked suffix.
    """
    if not account_suffix:
        raise SystemExit("ERROR: Could not determine account suffix from the Schwab lots title line.")

    from openpyxl import load_workbook

    securities: dict[str, Security] = {}
    current_security: Optional[Security] = None
    current_account_matches = False
    matched_account_name: Optional[str] = None

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]

        for row in ws.iter_rows(values_only=True):
            cells = list(row[:7])
            while len(cells) < 7:
                cells.append(None)

            name = str(cells[1]).strip() if cells[1] is not None else ""
            ticker = str(cells[2]).strip().upper() if cells[2] is not None else ""
            quote = parse_number(cells[3])
            shares = parse_number(cells[4])
            market_value = parse_number(cells[5])
            cost_basis = parse_number(cells[6])

            # Quicken reports cash as a row named "Cash" without a ticker symbol.
            if current_account_matches and name.lower() == "cash":
                securities[CASH_SYMBOL] = Security(
                    name="Cash",
                    ticker=CASH_SYMBOL,
                    shares=0.0,
                    market_value=market_value or cost_basis,
                    cost_basis=cost_basis,
                )
                current_security = None
                continue

            # Account summary/header row, e.g. "Estate 3993".
            if name and not ticker and cells[4] is None:
                digits = "".join(ch for ch in name if ch.isdigit())
                current_account_matches = bool(digits) and digits.endswith(account_suffix)
                if current_account_matches:
                    matched_account_name = name
                current_security = None
                continue

            if not current_account_matches or not name:
                continue

            if name.startswith("Lot"):
                if current_security is not None:
                    current_security.lots.append(
                        Lot(
                            name=name,
                            quote=quote,
                            shares=shares,
                            market_value=market_value,
                            cost_basis=cost_basis,
                        )
                    )
                continue

            current_security = Security(
                name=name,
                ticker=ticker,
                quote=quote,
                shares=shares,
                market_value=market_value,
                cost_basis=cost_basis,
            )
            if ticker:
                securities[ticker] = current_security
    finally:
        wb.close()

    if not matched_account_name:
        raise SystemExit(
            "ERROR: Could not find a Quicken account ending in Schwab account suffix "
            f"{account_suffix}."
        )

    return AccountSecurities(securities=securities, account_name=matched_account_name)


def quicken_lots_by_date(security: Security) -> dict[str, dict[str, float]]:
    lots: dict[str, dict[str, float]] = {}
    for lot in security.lots:
        if not lot.name.startswith("Lot "):
            continue
        date_text = normalize_date(lot.name[4:].strip())
        if not date_text:
            continue
        add_lot(lots, date_text, lot.shares, lot.cost_basis)
    return lots


def compare_symbol(
    ticker: str,
    schwab_lots: dict[str, dict[str, float]],
    quicken_security: Optional[Security],
) -> SymbolComparison:
    if quicken_security is None:
        quicken_security = Security(name=ticker, ticker=ticker)

    comparisons: list[LotComparison] = []

    if ticker == CASH_SYMBOL:
        quicken_lots = {}
    else:
        quicken_lots = quicken_lots_by_date(quicken_security)

        for date_text in sorted(set(schwab_lots) | set(quicken_lots), key=sortable_date):
            schwab_data = schwab_lots.get(date_text, {})
            quicken_data = quicken_lots.get(date_text, {})

            schwab_shares = schwab_data.get("shares", 0.0)
            schwab_cost = schwab_data.get("cost_basis", 0.0)
            quicken_shares = quicken_data.get("shares", 0.0)
            quicken_cost = quicken_data.get("cost_basis", 0.0)

            if date_text in schwab_lots and date_text not in quicken_lots:
                status = "missing_in_quicken"
            elif date_text not in schwab_lots and date_text in quicken_lots:
                status = "missing_in_schwab"
            elif abs(schwab_shares - quicken_shares) > 0.01 or abs(schwab_cost - quicken_cost) > 0.01:
                status = "mismatch"
            else:
                status = "match"

            comparisons.append(
                LotComparison(
                    date=date_text,
                    schwab_shares=schwab_shares,
                    quicken_shares=quicken_shares,
                    schwab_cost_basis=schwab_cost,
                    quicken_cost_basis=quicken_cost,
                    status=status,
                )
            )

    schwab_total_shares = sum(lot["shares"] for lot in schwab_lots.values())
    schwab_total_cost = sum(lot["cost_basis"] for lot in schwab_lots.values())
    schwab_total_market_value = sum(lot["market_value"] for lot in schwab_lots.values())
    schwab_lot_count = 0 if ticker == CASH_SYMBOL else len(schwab_lots)
    quicken_lot_count = 0 if ticker == CASH_SYMBOL else len(quicken_lots)

    return SymbolComparison(
        symbol=ticker,
        name=quicken_security.name,
        schwab_shares=schwab_total_shares,
        schwab_cost_basis=schwab_total_cost,
        quicken_shares=quicken_security.shares,
        quicken_cost_basis=quicken_security.cost_basis,
        schwab_lots=schwab_lot_count,
        quicken_lots=quicken_lot_count,
        market_value=schwab_total_market_value,
        lot_comparisons=comparisons,
    )


def format_float(value: float, precision: int = 2) -> str:
    return f"{value:,.{precision}f}"


def format_currency(value: float) -> str:
    return f"${value:,.2f}"


def comparison_flag(
    schwab_shares: float,
    quicken_shares: float,
    schwab_cost_basis: float,
    quicken_cost_basis: float,
) -> str:
    shares_differ = abs(schwab_shares - quicken_shares) > 0.01
    cost_basis_differs = abs(schwab_cost_basis - quicken_cost_basis) > 0.01
    return "*" if shares_differ or cost_basis_differs else ""


def symbol_table_rows(comparisons: list[SymbolComparison]) -> tuple[list[str], list[list[str]]]:
    headers = [
        "Symbol",
        "Flag",
        "Schwab Shares",
        "Quicken Shares",
        "Schwab Cost Basis",
        "Quicken Cost Basis",
        "Schwab Lots",
        "Quicken Lots",
        "Market Value",
    ]
    rows = [
        [
            comparison.symbol,
            comparison_flag(
                comparison.schwab_shares,
                comparison.quicken_shares,
                comparison.schwab_cost_basis,
                comparison.quicken_cost_basis,
            ),
            format_float(comparison.schwab_shares, 4),
            format_float(comparison.quicken_shares, 4),
            format_currency(comparison.schwab_cost_basis),
            format_currency(comparison.quicken_cost_basis),
            str(comparison.schwab_lots),
            str(comparison.quicken_lots),
            format_currency(comparison.market_value),
        ]
        for comparison in comparisons
    ]
    total_schwab_shares = sum(comparison.schwab_shares for comparison in comparisons)
    total_quicken_shares = sum(comparison.quicken_shares for comparison in comparisons)
    total_schwab_cost_basis = sum(comparison.schwab_cost_basis for comparison in comparisons)
    total_quicken_cost_basis = sum(comparison.quicken_cost_basis for comparison in comparisons)
    total_schwab_lots = sum(comparison.schwab_lots for comparison in comparisons)
    total_quicken_lots = sum(comparison.quicken_lots for comparison in comparisons)
    rows.append(
        [
            "TOTAL",
            comparison_flag(
                total_schwab_shares,
                total_quicken_shares,
                total_schwab_cost_basis,
                total_quicken_cost_basis,
            ),
            format_float(total_schwab_shares, 4),
            format_float(total_quicken_shares, 4),
            format_currency(total_schwab_cost_basis),
            format_currency(total_quicken_cost_basis),
            str(total_schwab_lots),
            str(total_quicken_lots),
            format_currency(sum(comparison.market_value for comparison in comparisons)),
        ]
    )
    return headers, rows


def print_symbol_table(comparisons: list[SymbolComparison]) -> None:
    headers, rows = symbol_table_rows(comparisons)
    left_aligned_columns = {0, 1}
    widths = [
        max(len(header), *(len(row[index]) for row in rows))
        for index, header in enumerate(headers)
    ]

    header = "  ".join(
        value.ljust(widths[index]) if index in left_aligned_columns else value.rjust(widths[index])
        for index, value in enumerate(headers)
    )
    separator = "  ".join("-" * width for width in widths)

    print(header)
    print(separator)
    for index, row in enumerate(rows):
        if index == len(rows) - 1:
            print(separator)
        print(
            "  ".join(
                value.ljust(widths[index]) if index in left_aligned_columns else value.rjust(widths[index])
                for index, value in enumerate(row)
            )
        )


def print_markdown_symbol_table(comparisons: list[SymbolComparison], account_name: Optional[str]) -> None:
    headers, rows = symbol_table_rows(comparisons)
    print(f"# {account_name or 'Quicken Account'}")
    print()
    print("| " + " | ".join(headers) + " |")
    print("| " + " | ".join("---" for _ in headers) + " |")
    for row in rows:
        print("| " + " | ".join(row) + " |")


def print_csv_symbol_table(comparisons: list[SymbolComparison]) -> None:
    headers, rows = symbol_table_rows(comparisons)
    writer = csv.writer(sys.stdout)
    writer.writerow(headers)
    writer.writerows(rows)


def print_html_symbol_table(comparisons: list[SymbolComparison], account_name: Optional[str]) -> None:
    headers, rows = symbol_table_rows(comparisons)
    title = account_name or "Quicken Account"

    print("<!doctype html>")
    print('<html lang="en">')
    print("<head>")
    print('  <meta charset="utf-8">')
    print('  <meta name="viewport" content="width=device-width, initial-scale=1">')
    print(f"  <title>{html.escape(title)}</title>")
    print("  <style>")
    print("    body { font-family: system-ui, -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; margin: 24px; }")
    print("    table { border-collapse: collapse; width: auto; max-width: none; }")
    print("    th, td { border: 1px solid #d0d7de; padding: 6px 10px; white-space: nowrap; }")
    print("    th { background: #f6f8fa; text-align: right; }")
    print("    th:first-child, td:first-child, th:nth-child(2), td:nth-child(2) { text-align: left; }")
    print("    td { text-align: right; }")
    print("    tr.flagged { background: #fff8c5; }")
    print("    tbody tr:last-child { font-weight: 700; }")
    print("  </style>")
    print("</head>")
    print("<body>")
    print(f"  <h1>{html.escape(title)}</h1>")
    print("  <table>")
    print("    <thead>")
    print("      <tr>")
    for header in headers:
        print(f"        <th>{html.escape(header)}</th>")
    print("      </tr>")
    print("    </thead>")
    print("    <tbody>")
    for row in rows:
        row_class = ' class="flagged"' if row[1] else ""
        print(f"      <tr{row_class}>")
        for value in row:
            print(f"        <td>{html.escape(value)}</td>")
        print("      </tr>")
    print("    </tbody>")
    print("  </table>")
    print("</body>")
    print("</html>")


def print_symbol_report(
    comparisons: list[SymbolComparison],
    output_format: str,
    account_name: Optional[str],
) -> None:
    if output_format == "text":
        print_symbol_table(comparisons)
    elif output_format == "markdown":
        print_markdown_symbol_table(comparisons, account_name)
    elif output_format == "csv":
        print_csv_symbol_table(comparisons)
    elif output_format == "html":
        print_html_symbol_table(comparisons, account_name)
    else:
        raise ValueError(f"Unsupported output format: {output_format}")


def print_lot_discrepancies(comparisons: list[SymbolComparison], output=sys.stdout) -> None:
    printed_header = False
    for symbol_comparison in comparisons:
        discrepancies = [
            comparison
            for comparison in symbol_comparison.lot_comparisons
            if comparison.status != "match"
        ]
        if not discrepancies:
            continue

        if not printed_header:
            print("\nLot Discrepancies:", file=output)
            printed_header = True
        print(f"\n{symbol_comparison.symbol}:", file=output)
        for discrepancy in discrepancies:
            print(discrepancy, file=output)

    if not printed_header:
        print("\nNo lot discrepancies.", file=output)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Compare Schwab lot data to the latest matching Quicken lots workbook."
    )
    parser.add_argument(
        "--show-discrepancies",
        action="store_true",
        help="display individual lot discrepancy details",
    )
    parser.add_argument(
        "--format",
        choices=("text", "markdown", "csv", "html"),
        default="text",
        help="output table format (default: text)",
    )
    args = parser.parse_args()
    status_output = sys.stderr if args.format in {"csv", "markdown", "html"} else sys.stdout

    schwab_file = find_schwab_lot_file()
    quicken_file = find_latest_quicken_lots_file()

    print(f"\n=== Parsing Schwab lots: {schwab_file} ===", file=status_output)
    schwab = load_schwab_lots(schwab_file)
    print(f"Found {len(schwab.lots_by_symbol)} symbols in Schwab lots", file=status_output)

    quicken = parse_quicken_lots(quicken_file, schwab.account_suffix)

    comparisons: list[SymbolComparison] = []
    for ticker in sorted(set(schwab.lots_by_symbol) | set(quicken.securities)):
        comparisons.append(
            compare_symbol(
                ticker,
                schwab.lots_by_symbol.get(ticker, {}),
                quicken.securities.get(ticker),
            )
        )

    if args.format not in {"markdown", "html"}:
        print(f"\nUsing Quicken account: {quicken.account_name}", file=status_output)
    print_symbol_report(comparisons, args.format, quicken.account_name)

    if args.show_discrepancies:
        print_lot_discrepancies(comparisons, output=status_output)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
