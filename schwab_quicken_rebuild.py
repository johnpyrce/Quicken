#!/usr/bin/env python3
"""
schwab_quicken_rebuild.py

Does ALL the upgrades:

1) Reads ONE Schwab "Lot Details" CSV per security from ./lots/   (no parameter)
2) Groups lots into 4 buckets:
      VERY_OLD  : acq < 2020-01-01
      OLD       : 2020-01-01 <= acq < 2023-01-01
      LONG_RECENT: acq >= 2023-01-01 and acq < (as_of_date - 365 days)
      SHORT_TERM: acq >= (as_of_date - 365 days)
   - as_of_date is parsed from each file’s title line (falls back to today if missing)

3) Outputs:
   - quicken_addshares_checklist.csv     (what you key into Quicken: Add Shares)
   - quicken_security_subtotals.csv      (per-security totals + counts)
   - schwab_quicken_validation_report.csv (diff vs Schwab positions summary, if found)

4) Validation (optional but automatic):
   - If a Schwab "Positions/Summary" CSV exists in the current folder, named:
        schwab_positions.csv
     …it will compare per-symbol Shares + Cost Basis totals and write a report.
   - If not found, it still writes the other files and notes validation was skipped.

How to run:
  Put your Schwab per-security lot CSVs in ./lots/*.csv
  (Optional) Put Schwab positions export in ./schwab_positions.csv
  Run: python schwab_quicken_rebuild.py
"""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Set


# -------------------------
# Fixed locations / default outputs (per your request)
# -------------------------

LOTS_DIR = Path("lots")  # required name; no parameter
POSITIONS_SUMMARY_FILE = Path("Quicken_Lots.xlsx")  # optional; auto-used if present

OUT_CHECKLIST = Path("quicken_addshares_checklist.csv")
OUT_SUBTOTALS = Path("quicken_security_subtotals.csv")
OUT_VALIDATION = Path("schwab_quicken_validation_report.csv")

MEMO_TEXT = "Schwab rebuild – grouped lots"


# -------------------------
# Schwab lot-details column names (from your AAPL file)
# -------------------------
DATE_COL = "Open Date"
SHARES_COL = "Quantity"
COST_COL = "Cost Basis"


# -------------------------
# Bucket definitions
# -------------------------
BREAK_1 = date(2020, 1, 1)
BREAK_2 = date(2023, 1, 1)

BUCKET_NAMES = ["VERY_OLD", "OLD", "LONG_RECENT", "SHORT_TERM"]

# Quicken acquisition dates used for the aggregated Add Shares lines.
# Note: SHORT_TERM uses the file’s as_of_date (so it stays obviously short-term).
QUICKEN_DATES_FIXED = {
    "VERY_OLD": date(2019, 12, 31),
    "OLD": date(2022, 12, 31),
    "LONG_RECENT": date(2024, 12, 31),
    # SHORT_TERM: dynamic per file
}


# -------------------------
# Parsing helpers
# -------------------------

def parse_date_any(s: str) -> date:
    s = (s or "").strip()
    if not s:
        raise ValueError("Empty date")
    fmts = ["%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y", "%Y/%m/%d"]
    for f in fmts:
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            pass
    # last resort: try extracting MM/DD/YYYY
    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return date(int(m.group(3)), int(m.group(1)), int(m.group(2)))
    raise ValueError(f"Unrecognized date format: {s!r}")


def parse_decimal_any(s) -> Decimal:
    if s is None:
        raise ValueError("Empty numeric field")

    if isinstance(s, Decimal):
        return s

    if isinstance(s, (int, float)):
        return Decimal(str(s))

    s = str(s).strip()
    if not s:
        raise ValueError("Empty numeric field")
    
    # Handle placeholder values from Schwab exports
    if s in ("--", "N/A", "n/a", "*"):
        return Decimal(0)

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()

    s = s.replace("$", "").replace(",", "").replace("*", "").strip()
    try:
        v = Decimal(s)
    except InvalidOperation:
        raise ValueError(f"Unrecognized numeric format: {s!r}")

    return -v if neg else v


def mmddyyyy(d: date) -> str:
    return d.strftime("%m/%d/%Y")


# -------------------------
# Data
# -------------------------

@dataclass(frozen=True)
class Lot:
    symbol: str
    acq_date: date
    shares: Decimal
    cost: Decimal


# -------------------------
# Schwab lot CSV parsing (your per-security format)
# -------------------------

def extract_symbol_from_title(title_line: str) -> str:
    """
    Example: "AAPL Lot Details for ... as of 12:18 PM ET, 12/23/2025"
    """
    title_line = (title_line or "").strip()
    m = re.match(r"^\s*([A-Z0-9.\-]+)\s+Lot Details\b", title_line)
    if m:
        return m.group(1).upper()
    # fallback: first token
    tok = title_line.split(" ", 1)[0] if title_line else ""
    return tok.upper() if tok else "UNKNOWN"


def extract_as_of_date_from_title(title_line: str) -> Optional[date]:
    """
    Tries to find "... as of ..., 12/23/2025" or any MM/DD/YYYY in the title line.
    """
    title_line = (title_line or "").strip()
    m = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", title_line)
    if m:
        try:
            return parse_date_any(m.group(1))
        except Exception:
            return None
    return None


def extract_account_suffix_from_title(title_line: str) -> Optional[str]:
    """
    Example: "AAPL Lot Details for  ...993 as of 01:13 PM ET, 04/25/2026"
    Returns the visible masked account suffix, e.g. "993".
    """
    title_line = (title_line or "").strip()
    m = re.search(r"for\s+\.\.\.(\d{3,4})\b", title_line, re.IGNORECASE)
    if m:
        return m.group(1)
    return None


def slugify_account_name(name: str) -> str:
    slug = re.sub(r"[^A-Za-z0-9]+", "_", (name or "").strip()).strip("_").lower()
    return slug or "account"


def find_header_row(rows: List[List[str]]) -> int:
    """Find the row containing the required column headers (in any order)."""
    for i, r in enumerate(rows):
        # Check if this row contains all required column names
        row_stripped = [c.strip() for c in r]
        if DATE_COL in row_stripped and SHARES_COL in row_stripped and COST_COL in row_stripped:
            return i
    # Debug: show what we found
    print(f"  DEBUG: Could not find header. First 5 rows:")
    for i, r in enumerate(rows[:5]):
        print(f"    Row {i}: {r[:5] if len(r) >= 5 else r}")
    raise ValueError(f"Could not find header row containing '{DATE_COL}', '{SHARES_COL}', and '{COST_COL}'.")


def load_lots_file(path: Path) -> Tuple[List[Lot], date]:
    """
    Returns (lots, as_of_date_used)
    """
    # Read entire CSV with csv.reader to handle the title row + header row
    text = path.read_text(encoding="utf-8-sig")
    rows = list(csv.reader(text.splitlines()))
    if not rows:
        return ([], date.today())

    title_line = rows[0][0] if rows[0] else ""
    symbol = extract_symbol_from_title(title_line)
    as_of = extract_as_of_date_from_title(title_line) or date.today()

    hdr_idx = find_header_row(rows)
    header = rows[hdr_idx]
    idx = {h.strip(): i for i, h in enumerate(header)}

    for needed in (DATE_COL, SHARES_COL, COST_COL):
        if needed not in idx:
            raise ValueError(f"{path.name}: Missing expected column {needed!r}. Found: {header}")

    lots: List[Lot] = []
    for r in rows[hdr_idx + 1 :]:
        if not r or all((c or "").strip() == "" for c in r):
            continue

        d_raw = r[idx[DATE_COL]].strip() if idx[DATE_COL] < len(r) else ""
        if not d_raw:
            continue
        
        # Skip summary/total rows
        if d_raw.lower() in ("total", "totals", "grand total", "summary"):
            continue

        try:
            acq = parse_date_any(d_raw)
        except ValueError:
            # Skip rows with invalid dates (likely summary rows)
            continue
            
        shares_raw = r[idx[SHARES_COL]] if idx[SHARES_COL] < len(r) else ""
        cost_raw = r[idx[COST_COL]] if idx[COST_COL] < len(r) else ""

        shares = parse_decimal_any(shares_raw)
        cost = parse_decimal_any(cost_raw)

        if shares == 0:
            continue

        lots.append(Lot(symbol=symbol, acq_date=acq, shares=shares, cost=cost))

    return (lots, as_of)


def load_cash_file(path: Path) -> Decimal:
    text = path.read_text(encoding="utf-8-sig")
    rows = list(csv.reader(text.splitlines()))
    if len(rows) < 4:
        return Decimal("0")

    header = rows[2]
    idx = {h.strip(): i for i, h in enumerate(header)}
    cost_idx = idx.get("Cost Basis")
    if cost_idx is None:
        return Decimal("0")

    value_row = rows[3]
    cost_raw = value_row[cost_idx] if cost_idx < len(value_row) else "0"
    return parse_decimal_any(cost_raw)


# -------------------------
# Bucketing
# -------------------------

def bucket_name_for(acq: date, as_of: date) -> str:
    cutoff_short = as_of - timedelta(days=365)

    if acq < BREAK_1:
        return "VERY_OLD"
    if acq < BREAK_2:
        return "OLD"
    if acq < cutoff_short:
        return "LONG_RECENT"
    return "SHORT_TERM"


def quicken_bucket_date(bucket: str, as_of: date) -> date:
    if bucket == "SHORT_TERM":
        return as_of
    return QUICKEN_DATES_FIXED[bucket]


# -------------------------
# Aggregation
# -------------------------

@dataclass
class AggRow:
    symbol: str
    bucket: str
    shares: Decimal
    cost: Decimal
    quicken_date: date
    lots_count: int


def aggregate_all(files: List[Path]) -> Tuple[List[AggRow], Dict[str, Dict[str, Decimal]], Dict[str, Dict[str, int]]]:
    """
    Returns:
      - list of aggregated bucket rows (for checklist)
      - per-security totals {"shares":..., "cost":...}
      - per-security counts {"lot_files":..., "lots":..., "bucket_rows":...}
    """
    agg: Dict[Tuple[str, str], Dict[str, Decimal]] = {}
    counts: Dict[Tuple[str, str], int] = {}
    quick_dates: Dict[Tuple[str, str], date] = {}
    per_symbol_totals: Dict[str, Dict[str, Decimal]] = {}
    per_symbol_counts: Dict[str, Dict[str, int]] = {}

    for f in files:
        if f.stem.lower() == "cash":
            cash_cost = load_cash_file(f)
            sym = "CASH"
            per_symbol_counts.setdefault(sym, {"lot_files": 0, "lots": 0, "bucket_rows": 0})
            per_symbol_counts[sym]["lot_files"] += 1
            per_symbol_counts[sym]["lots"] += 1
            per_symbol_totals[sym] = {"shares": Decimal("0"), "cost": cash_cost}
            continue

        lots, as_of = load_lots_file(f)
        if not lots:
            continue

        sym = lots[0].symbol if lots else "UNKNOWN"
        per_symbol_counts.setdefault(sym, {"lot_files": 0, "lots": 0, "bucket_rows": 0})
        per_symbol_counts[sym]["lot_files"] += 1

        per_symbol_totals.setdefault(sym, {"shares": Decimal("0"), "cost": Decimal("0")})

        for lot in lots:
            b = bucket_name_for(lot.acq_date, as_of)
            key = (lot.symbol, b)
            agg.setdefault(key, {"shares": Decimal("0"), "cost": Decimal("0")})
            counts[key] = counts.get(key, 0) + 1

            agg[key]["shares"] += lot.shares
            agg[key]["cost"] += lot.cost

            # Remember the quicken date per (symbol,bucket). For SHORT_TERM, we want the *latest* as_of date encountered.
            qd = quicken_bucket_date(b, as_of)
            prev = quick_dates.get(key)
            if prev is None or (b == "SHORT_TERM" and qd > prev):
                quick_dates[key] = qd

            per_symbol_totals[lot.symbol]["shares"] += lot.shares
            per_symbol_totals[lot.symbol]["cost"] += lot.cost
            per_symbol_counts[lot.symbol]["lots"] += 1

    # Build AggRow list
    out_rows: List[AggRow] = []
    for (sym, b), sums in agg.items():
        sh = sums["shares"]
        if sh == 0:
            continue
        out_rows.append(
            AggRow(
                symbol=sym,
                bucket=b,
                shares=sh,
                cost=sums["cost"],
                quicken_date=quick_dates[(sym, b)],
                lots_count=counts.get((sym, b), 0),
            )
        )

    # count bucket rows per symbol
    for r in out_rows:
        per_symbol_counts.setdefault(r.symbol, {"lot_files": 0, "lots": 0, "bucket_rows": 0})
        per_symbol_counts[r.symbol]["bucket_rows"] += 1

    # Sort rows for easy manual entry
    bucket_order = {name: i for i, name in enumerate(BUCKET_NAMES)}
    out_rows.sort(key=lambda r: (r.symbol, bucket_order.get(r.bucket, 99)))

    return out_rows, per_symbol_totals, per_symbol_counts


# -------------------------
# Positions summary validation
# -------------------------

def detect_account_suffix_from_lot_files(paths: List[Path]) -> Optional[str]:
    suffixes: Set[str] = set()

    for path in paths:
        text = path.read_text(encoding="utf-8-sig")
        rows = list(csv.reader(text.splitlines()))
        if not rows:
            continue
        title_line = rows[0][0] if rows[0] else ""
        suffix = extract_account_suffix_from_title(title_line)
        if suffix:
            suffixes.add(suffix)

    if not suffixes:
        return None
    if len(suffixes) > 1:
        raise ValueError(f"Lot files span multiple account suffixes: {sorted(suffixes)}")
    return next(iter(suffixes))


def build_output_paths(account_name: Optional[str]) -> Tuple[Path, Path, Path]:
    if not account_name:
        return OUT_CHECKLIST, OUT_SUBTOTALS, OUT_VALIDATION

    slug = slugify_account_name(account_name)
    return (
        Path(f"{slug}_quicken_addshares_checklist.csv"),
        Path(f"{slug}_quicken_security_subtotals.csv"),
        Path(f"{slug}_schwab_quicken_validation_report.csv"),
    )


def load_positions_summary(path: Path) -> Tuple[Dict[str, Dict[str, Decimal]], Optional[str]]:
    """
    Returns per-symbol totals from the Quicken_Lots workbook for the account
    matching the lot file account suffix:
      { "AAPL": {"shares":..., "cost":...}, ... }
    """
    out: Dict[str, Dict[str, Decimal]] = {}

    account_suffix = detect_account_suffix_from_lot_files(sorted(LOTS_DIR.glob("*.csv")))
    if not account_suffix:
        raise ValueError("Could not determine account suffix from lot file titles.")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    current_account_matches = False
    matched_account_name: Optional[str] = None

    for row in ws.iter_rows(values_only=True):
        cells = list(row[:7])
        while len(cells) < 7:
            cells.append(None)

        name = str(cells[1]).strip() if cells[1] is not None else ""
        symbol = str(cells[2]).strip().upper() if cells[2] is not None else ""
        shares_val = cells[4]
        cost_val = cells[6]

        # Quicken reports cash as a row named "Cash" without a ticker symbol.
        if current_account_matches and name.lower() == "cash":
            out["CASH"] = {
                "shares": Decimal("0"),
                "cost": parse_decimal_any(cost_val),
            }
            continue

        # Account summary/header row, e.g. "Estate 3993"
        if name and not symbol and shares_val is None:
            digits = "".join(ch for ch in name if ch.isdigit())
            current_account_matches = bool(digits) and digits.endswith(account_suffix)
            if current_account_matches:
                matched_account_name = name
            continue

        if not current_account_matches:
            continue

        if not symbol:
            continue

        # Keep only per-security summary rows, not lot-detail rows.
        if shares_val is None:
            continue

        if symbol in ("TOTAL", "TOTALS"):
            continue

        out[symbol] = {
            "shares": parse_decimal_any(shares_val),
            "cost": parse_decimal_any(cost_val),
        }

    wb.close()
    return out, matched_account_name


def write_validation_report(
    quicken_pos: Dict[str, Dict[str, Decimal]],
    schwab_lot_totals: Dict[str, Dict[str, Decimal]],
    output_path: Path,
) -> None:
    """
    Writes per-symbol diffs and a status column.
    """
    symbols = sorted(set(quicken_pos.keys()) | set(schwab_lot_totals.keys()))
    rows = []
    for sym in symbols:
        q = quicken_pos.get(sym)
        s = schwab_lot_totals.get(sym)

        q_sh = q["shares"] if q else None
        q_cost = q["cost"] if q else None
        s_sh = s["shares"] if s else None
        s_cost = s["cost"] if s else None

        sh_diff = (s_sh - q_sh) if (q_sh is not None and s_sh is not None) else None
        cost_diff = (s_cost - q_cost) if (q_cost is not None and s_cost is not None) else None

        status = "OK"
        if q is None:
            status = "MISSING_IN_QUICKEN"
        elif s is None:
            status = "MISSING_IN_SCHWAB_LOTS"
        else:
            # Tolerances: shares exact (to 0.0001), cost within $0.01
            if sh_diff is not None and cost_diff is not None:
                if abs(sh_diff) > Decimal("0.0001") or abs(cost_diff) > Decimal("0.01"):
                    status = "MISMATCH"

        rows.append({
            "Symbol": sym,
            "QuickenShares": "" if q_sh is None else str(q_sh),
            "SchwabShares": "" if s_sh is None else str(s_sh),
            "ShareDiff(Schwab-Quicken)": "" if sh_diff is None else str(sh_diff),
            "QuickenCostBasis": "" if q_cost is None else str(q_cost),
            "SchwabCostBasis": "" if s_cost is None else str(s_cost),
            "CostDiff(Schwab-Quicken)": "" if cost_diff is None else str(cost_diff),
            "Status": status,
        })

    with output_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()) if rows else [
            "Symbol","QuickenShares","SchwabShares","ShareDiff(Schwab-Quicken)","QuickenCostBasis",
            "SchwabCostBasis","CostDiff(Schwab-Quicken)","Status"
        ])
        w.writeheader()
        w.writerows(rows)


# -------------------------
# Writers
# -------------------------

def write_checklist(rows: List[AggRow], output_path: Path) -> None:
    with output_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["Symbol", "Bucket", "Shares", "TotalCost", "AcqDateForQuicken", "LotsInBucket", "Memo"],
        )
        w.writeheader()
        for r in rows:
            w.writerow({
                "Symbol": r.symbol,
                "Bucket": r.bucket,
                "Shares": str(r.shares),
                "TotalCost": str(r.cost),
                "AcqDateForQuicken": mmddyyyy(r.quicken_date),
                "LotsInBucket": r.lots_count,
                "Memo": MEMO_TEXT,
            })


def write_subtotals(
    totals: Dict[str, Dict[str, Decimal]],
    counts: Dict[str, Dict[str, int]],
    output_path: Path,
) -> None:
    syms = sorted(totals.keys())
    with output_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["Symbol", "TotalShares", "TotalCostBasis", "LotFilesRead", "LotsRead", "BucketRowsEmitted"],
        )
        w.writeheader()
        for sym in syms:
            t = totals[sym]
            c = counts.get(sym, {"lot_files": 0, "lots": 0, "bucket_rows": 0})
            w.writerow({
                "Symbol": sym,
                "TotalShares": str(t["shares"]),
                "TotalCostBasis": str(t["cost"]),
                "LotFilesRead": c["lot_files"],
                "LotsRead": c["lots"],
                "BucketRowsEmitted": c["bucket_rows"],
            })


# -------------------------
# Main
# -------------------------

def main() -> None:
    if not LOTS_DIR.exists() or not LOTS_DIR.is_dir():
        raise SystemExit("Expected a subdirectory named 'lots' in the current folder.")

    lot_files = sorted(LOTS_DIR.glob("*.csv"))
    if not lot_files:
        raise SystemExit("No CSV files found in ./lots/")

    agg_rows, per_symbol_totals, per_symbol_counts = aggregate_all(lot_files)
    account_suffix = detect_account_suffix_from_lot_files(lot_files)
    account_name = f"account_{account_suffix}" if account_suffix else None

    if POSITIONS_SUMMARY_FILE.exists():
        _, workbook_account_name = load_positions_summary(POSITIONS_SUMMARY_FILE)
        if workbook_account_name:
            account_name = workbook_account_name

    checklist_path, subtotals_path, validation_path = build_output_paths(account_name)

    write_checklist(agg_rows, checklist_path)
    write_subtotals(per_symbol_totals, per_symbol_counts, subtotals_path)

    # Optional validation
    if POSITIONS_SUMMARY_FILE.exists():
        schwab_pos, workbook_account_name = load_positions_summary(POSITIONS_SUMMARY_FILE)
        if workbook_account_name:
            checklist_path, subtotals_path, validation_path = build_output_paths(workbook_account_name)
        write_validation_report(schwab_pos, per_symbol_totals, validation_path)
        validation_msg = f"Validation written: {validation_path}"
    else:
        # Still emit an empty-ish report that says skipped (handy reminder)
        with validation_path.open("w", newline="", encoding="utf-8") as f:
            f.write("Validation skipped: file Quicken_Lots.xlsx not found in current folder.\n")
        validation_msg = "Validation skipped (no Quicken_Lots.xlsx); stub report written."

    print(f"Read lot files: {len(lot_files)} from ./{LOTS_DIR}/")
    print(f"Wrote checklist: {checklist_path} ({len(agg_rows)} Add-Shares lines)")
    print(f"Wrote subtotals: {subtotals_path} ({len(per_symbol_totals)} securities)")
    print(validation_msg)


if __name__ == "__main__":
    main()
